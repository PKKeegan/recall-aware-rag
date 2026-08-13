"""
load_dataset.py

Utilities for loading the Recall-Aware RAG
evaluation dataset from Excel.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import openpyxl


DEFAULT_DATASET_PATH = Path(
    "data/evaluation/evaluation_dataset.xlsx"
)


REQUIRED_COLUMNS = {
    "question_id",
    "question",
    "category",
    "difficulty",
    "gold_answer",
    "expected_behavior",
    "evidence_source",
    "evidence_location",
    "evidence_summary",
    "support_level",
    "reason_for_abstention",
}


def load_evaluation_dataset(
    dataset_path: str | Path = DEFAULT_DATASET_PATH,
) -> List[Dict[str, str]]:
    """
    Load the evaluation dataset from an Excel workbook.

    Parameters
    ----------
    dataset_path:
        Path to the .xlsx evaluation dataset.

    Returns
    -------
    List[Dict[str, str]]
        One dictionary per evaluation question.
    """

    dataset_path = Path(dataset_path)

    if not dataset_path.exists():
        raise FileNotFoundError(
            f"Evaluation dataset not found: {dataset_path}"
        )

    if dataset_path.suffix.lower() != ".xlsx":
        raise ValueError(
            "Evaluation dataset must be an .xlsx file."
        )

    workbook = openpyxl.load_workbook(
        dataset_path,
        read_only=True,
        data_only=True,
    )

    if not workbook.sheetnames:
        raise ValueError(
            "The Excel workbook contains no worksheets."
        )

    # Use the first worksheet.
    worksheet = workbook[workbook.sheetnames[0]]

    rows = list(
        worksheet.iter_rows(
            values_only=True
        )
    )

    workbook.close()

    if not rows:
        raise ValueError(
            "The evaluation dataset is empty."
        )

    # ---------------------------------------------------------
    # Read header row
    # ---------------------------------------------------------

    headers = []

    for value in rows[0]:

        if value is None:
            headers.append("")
        else:
            headers.append(
                str(value).strip()
            )

    # Remove completely empty columns from the end.
    while headers and headers[-1] == "":
        headers.pop()

    if not headers:
        raise ValueError(
            "The Excel worksheet does not contain a header row."
        )

    # ---------------------------------------------------------
    # Validate required columns
    # ---------------------------------------------------------

    available_columns = set(headers)

    missing_columns = (
        REQUIRED_COLUMNS
        - available_columns
    )

    if missing_columns:

        raise ValueError(
            "Evaluation dataset is missing "
            f"required columns: "
            f"{sorted(missing_columns)}"
        )

    # ---------------------------------------------------------
    # Convert rows to dictionaries
    # ---------------------------------------------------------

    dataset: List[Dict[str, str]] = []

    for row_number, row in enumerate(
        rows[1:],
        start=2,
    ):

        # Skip completely empty rows.
        if all(
            value is None
            for value in row
        ):
            continue

        record: Dict[str, str] = {}

        for index, header in enumerate(headers):

            if index < len(row):
                value = row[index]
            else:
                value = None

            if value is None:
                value = ""

            record[header] = str(value).strip()

        # -----------------------------------------------------
        # Validate question ID
        # -----------------------------------------------------

        if not record["question_id"]:

            raise ValueError(
                f"Row {row_number} has no question_id."
            )

        # -----------------------------------------------------
        # Validate question
        # -----------------------------------------------------

        if not record["question"]:

            raise ValueError(
                f"Row {row_number} "
                f"({record['question_id']}) "
                "has no question."
            )

        dataset.append(record)

    if not dataset:
        raise ValueError(
            "No evaluation questions were found "
            "in the Excel workbook."
        )

    return dataset